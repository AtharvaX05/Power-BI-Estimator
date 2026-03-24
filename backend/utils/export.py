"""Export estimation results to Excel and PDF."""
import io
from typing import Optional, List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from backend.models.project import Project, ProjectVersion


def _compute_cost_rows(total_hours: float, roles: List[Dict]) -> List[Dict]:
    """Compute hours and cost per role."""
    result = []
    for r in roles:
        pct = float(r.get("percentage", 0))
        rate = float(r.get("rate", 0))
        hours = round(total_hours * pct / 100, 1)
        cost = round(hours * rate, 2)
        result.append({
            "role": r.get("role", ""),
            "percentage": pct,
            "hours": hours,
            "rate": rate,
            "cost": cost,
        })
    return result


def export_to_excel(
    project: Project, version: ProjectVersion,
    roles: Optional[List[Dict]] = None, currency: str = "",
) -> io.BytesIO:
    """Generate .xlsx with effort breakdown, optional cost, then assumptions at end."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Effort Estimation"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="1B7340", end_color="1B7340", fill_type="solid")
    sym = currency if currency else ""
    if sym == "₹":
        sym = "Rs. "

    row = 1
    ws.merge_cells("A1:D1")
    title_text = "Power BI Effort & Cost Estimation Report" if roles else "Power BI Effort Estimation Report"
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    row += 2

    # Project details
    details = [
        ("Project Name", project.name),
        ("Client", project.client_name),
        ("Description", project.description),
        ("Version", f"v{version.version_number}"),
        ("Timestamp", version.timestamp.strftime("%Y-%m-%d %H:%M UTC")),
        ("Total Effort", f"{version.outputs.total_effort_hours} hours ({version.outputs.total_effort_days} days)"),
    ]
    for label, val in details:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1
    row += 1

    # Module breakdown
    headers = ["Module", "Estimate (Hrs)", "Estimate (Days)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center"); c.border = thin_border
    row += 1
    for m in version.outputs.module_breakdown:
        ws.cell(row=row, column=1, value=m.module_name).border = thin_border
        ws.cell(row=row, column=2, value=m.computed_effort_hours).border = thin_border
        ws.cell(row=row, column=3, value=round(m.computed_effort_hours / 8, 1)).border = thin_border
        row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=version.outputs.total_effort_hours).font = Font(bold=True)
    ws.cell(row=row, column=3, value=version.outputs.total_effort_days).font = Font(bold=True)
    row += 2

    # Cost Estimation section (if roles provided)
    if roles:
        ws.cell(row=row, column=1, value="Cost Estimation").font = Font(bold=True, size=13, color="1B7340")
        row += 1
        cost_headers = ["Role", "% Efforts", "Hours", f"Rate/hr ({sym})" if sym else "Rate/hr", f"Cost ({sym})" if sym else "Cost"]
        for col, h in enumerate(cost_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font; c.fill = green_fill
            c.alignment = Alignment(horizontal="center"); c.border = thin_border
        row += 1

        cost_rows = _compute_cost_rows(version.outputs.total_effort_hours, roles)
        total_cost = 0.0
        for cr in cost_rows:
            ws.cell(row=row, column=1, value=cr["role"]).border = thin_border
            ws.cell(row=row, column=2, value=f'{cr["percentage"]}%').border = thin_border
            ws.cell(row=row, column=3, value=cr["hours"]).border = thin_border
            ws.cell(row=row, column=4, value=cr["rate"]).border = thin_border
            ws.cell(row=row, column=5, value=cr["cost"]).border = thin_border
            total_cost += cr["cost"]
            row += 1

        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value="100%").font = Font(bold=True)
        ws.cell(row=row, column=3, value=version.outputs.total_effort_hours).font = Font(bold=True)
        ws.cell(row=row, column=5, value=round(total_cost, 2)).font = Font(bold=True)
        row += 2
        ws.cell(row=row, column=1, value="Total Cost").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{sym}{total_cost:,.2f}" if sym else round(total_cost, 2)).font = Font(bold=True)
        row += 2

    # Assumptions (always at the end)
    ws.cell(row=row, column=1, value="Assumptions").font = section_font
    row += 5
    for idx, a in enumerate(version.outputs.assumptions, 1):
        ws.cell(row=row, column=1, value=f"{idx}. {a}")
        row += 1

    # Auto-width
    for col_cells in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col_cells:
            if hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_to_pdf(
    project: Project, version: ProjectVersion,
    roles: Optional[List[Dict]] = None, currency: str = "",
) -> io.BytesIO:
    """Generate PDF with effort breakdown, optional cost, then assumptions at end."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleBlue", parent=styles["Title"], textColor=colors.HexColor("#2F5496"))
    sym = currency if currency else ""
    if sym == "₹":
        sym = "Rs. "
    elements = []


    title_text = "Power BI Effort & Cost Estimation Report" if roles else "Power BI Effort Estimation Report"
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 10))

    # Project info
    info_data = [
        ["Project Name", project.name],
        ["Client", project.client_name],
        ["Description", project.description or "—"],
        ["Version", f"v{version.version_number}"],
        ["Date", version.timestamp.strftime("%Y-%m-%d %H:%M UTC")],
        ["Total Effort", f"{version.outputs.total_effort_hours} hours ({version.outputs.total_effort_days} days)"],
    ]
    info_table = Table(info_data, colWidths=[120, 350])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 16))

    # Module breakdown
    elements.append(Paragraph("Module-wise Breakdown", styles["Heading2"]))
    table_data = [["Module", "Estimate (Hrs)", "Estimate (Days)"]]
    for m in version.outputs.module_breakdown:
        table_data.append([m.module_name, str(m.computed_effort_hours),
                           str(round(m.computed_effort_hours / 8, 1))])
    table_data.append(["TOTAL", str(version.outputs.total_effort_hours), str(version.outputs.total_effort_days)])
    t = Table(table_data, colWidths=[250, 100, 100])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 16))

    # Cost Estimation section (if roles provided)
    if roles:
        elements.append(Paragraph("Cost Estimation", styles["Heading2"]))
        cost_rows = _compute_cost_rows(version.outputs.total_effort_hours, roles)
        total_cost = sum(cr["cost"] for cr in cost_rows)

        rate_hdr = f"Rate/hr ({sym})" if sym else "Rate/hr"
        cost_hdr = f"Cost ({sym})" if sym else "Cost"
        cost_table_data = [["Role", "% Efforts", "Hours", rate_hdr, cost_hdr]]
        for cr in cost_rows:
            cost_table_data.append([
                cr["role"], f'{cr["percentage"]}%', str(cr["hours"]),
                f'{sym}{cr["rate"]:.2f}', f'{sym}{cr["cost"]:,.2f}',
            ])
        cost_table_data.append(["TOTAL", "100%", str(version.outputs.total_effort_hours), "", f"{sym}{total_cost:,.2f}"])

        ct = Table(cost_table_data, colWidths=[130, 65, 65, 65, 90])
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B7340")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(ct)
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"<b>Total Cost: {sym}{total_cost:,.2f}</b>", styles["Normal"]))
        elements.append(Spacer(1, 16))

    # Assumptions (always at the end)
    elements.append(Paragraph("Assumptions", styles["Heading2"]))
    for idx, a in enumerate(version.outputs.assumptions, 1):
        elements.append(Paragraph(f"{idx}. {a}", styles["Normal"]))
    elements.append(Spacer(1, 8))

    doc.build(elements)
    buf.seek(0)
    return buf


# ── Legacy cost-specific exports (now delegate to unified functions) ──

def export_to_excel_with_cost(
    project: Project, version: ProjectVersion, roles: List[Dict],
    currency: str = "",
) -> io.BytesIO:
    return export_to_excel(project, version, roles=roles, currency=currency)


def export_to_pdf_with_cost(
    project: Project, version: ProjectVersion, roles: List[Dict],
    currency: str = "",
) -> io.BytesIO:
    return export_to_pdf(project, version, roles=roles, currency=currency)
